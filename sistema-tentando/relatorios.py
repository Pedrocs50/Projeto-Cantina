'''PARA GERAR OS RELATÓRIOS EM EXCEL'''

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from banco import BancoDeDados

class GeradorRelatorio:
    @staticmethod
    def gerar_relatorio_vendas():
        banco = BancoDeDados()
        conn = banco.conectar()
        cursor = conn.cursor()
        
        # Consulta SQL para pegar os dados de vendas, produtos e funcionários
        cursor.execute('''
        SELECT Vendas.id_venda, Produtos.nome, Vendas.quantidade_vendida, Vendas.data_venda, Funcionarios.nome
        FROM Vendas
        LEFT JOIN Produtos ON Vendas.id_produto = Produtos.id_produto
        LEFT JOIN Funcionarios ON Vendas.id_funcionario = Funcionarios.id_funcionario''')
        
        vendas = cursor.fetchall()
        conn.close()

        # Criando um DataFrame do pandas com os dados
        df = pd.DataFrame(vendas, columns=["ID Venda", "Produto", "Quantidade", "Data", "Funcionário"])

        # Gerar o arquivo Excel com os dados
        df.to_excel("relatorio_vendas.xlsx", index=False, sheet_name="Planinha 1")

        # Carregar o arquivo Excel com openpyxl para formatação
        workbook = load_workbook("relatorio_vendas.xlsx")
        worksheet = workbook.active

        # --- Formatação do cabeçalho ---
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:  # Aplica no cabeçalho (primeira linha)
            cell.font = header_font
            cell.fill = header_fill

        # --- Adicionar bordas em todas as células ---
        border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )
        
        for row in worksheet.iter_rows():  # Aplica em todas as células
            for cell in row:
                cell.border = border

        # --- Alinhar as células no centro ---
        alignment = Alignment(horizontal="center", vertical="center")
        
        for row in worksheet.iter_rows():  # Aplica no conteúdo de todas as células
            for cell in row:
                cell.alignment = alignment

        # --- Ajustar largura das colunas ---
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # Pega a letra da coluna (A, B, C, etc.)
            
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            
            adjusted_width = (max_length + 2)  # Ajusta a largura com um pequeno buffer
            worksheet.column_dimensions[column].width = adjusted_width

        # Salvar as alterações feitas na formatação
        workbook.save("relatorio_vendas_formatado.xlsx")

        print("Relatório de vendas gerado e formatado com sucesso: 'relatorio_vendas_formatado.xlsx'")

# Chamada do método para gerar o relatório
GeradorRelatorio.gerar_relatorio_vendas()

